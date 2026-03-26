import os
import re
import threading
import flet as ft
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def exportar_excel(caminho_pdf: str, com_diferenca: list, consolidado: list):
    wb = openpyxl.Workbook()

    cab_font = Font(bold=True, color="FFFFFF")
    cab_fill_dif = PatternFill("solid", fgColor="C62828")  # vermelho
    cab_fill_cons = PatternFill("solid", fgColor="2E7D32")  # verde
    centro = Alignment(horizontal="center")

    colunas = ["NF", "Crédito R$", "Débito R$", "Diferença R$"]

    for titulo, dados, fill in [
        ("Com Diferença", com_diferenca, cab_fill_dif),
        ("Consolidado", consolidado, cab_fill_cons),
    ]:
        ws = wb.create_sheet(titulo)

        for col, cab in enumerate(colunas, start=1):
            cell = ws.cell(row=1, column=col, value=cab)
            cell.font = cab_font
            cell.fill = fill
            cell.alignment = centro

        for linha, item in enumerate(dados, start=2):
            ws.cell(row=linha, column=1, value=item['nf'])
            ws.cell(row=linha, column=2, value=item['credito'])
            ws.cell(row=linha, column=3, value=item['debito'])
            ws.cell(row=linha, column=4, value=item['diferenca'])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    # Remove a aba vazia padrão
    del wb["Sheet"]

    # Salva na mesma pasta do PDF com sufixo _consolidado.xlsx
    base = os.path.splitext(caminho_pdf)[0]
    destino = f"{base}_consolidado.xlsx"
    wb.save(destino)
    return destino

# ── Processamento do PDF ─────────────────────────────────────────────────────


def processar_pdf(caminho_pdf):
    if not os.path.exists(caminho_pdf):
        return [], []

    dados_brutos = {}
    padrao_nota = re.compile(r"NF[e]?\s+(\d+)", re.IGNORECASE)
    padrao_valor = re.compile(r"(\d{1,3}(?:\.\d{3})*,\d{2})")

    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text()
            if not texto:
                continue
            for linha in texto.split('\n'):
                match_nota = padrao_nota.search(linha)
                if match_nota:
                    numero_nf = match_nota.group(1)
                    valores_texto = padrao_valor.findall(linha)
                    valores_float = [
                        float(v.replace('.', '').replace(',', '.'))
                        for v in valores_texto
                    ]
                    if numero_nf not in dados_brutos:
                        dados_brutos[numero_nf] = {'credito': 0.0, 'debito': 0.0}

                    linha_upper = linha.upper()
                    if "RECEBIMENTO" in linha_upper and valores_float:
                        dados_brutos[numero_nf]['credito'] += valores_float[0]
                    elif ("PRESTAÇÃO" in linha_upper or "NFE" in linha_upper) and valores_float:
                        dados_brutos[numero_nf]['debito'] += valores_float[0]

    com_diferenca = []
    consolidado = []

    for nf in sorted(dados_brutos.keys(), key=int):
        cred = dados_brutos[nf]['credito']
        deb = dados_brutos[nf]['debito']
        dif = round(cred - deb, 2)
        item = {'nf': nf, 'credito': cred, 'debito': deb, 'diferenca': dif}
        if dif != 0:
            com_diferenca.append(item)
        elif cred > 0:
            consolidado.append(item)

    return com_diferenca, consolidado


def formatar_valor(valor: float) -> str:
    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


# ── Componente de tabela paginada ────────────────────────────────────────────

class TabelaPaginada:
    LINHAS_POR_PAGINA = 50

    def __init__(self, page: ft.Page, dados: list, tem_diferenca: bool):
        self.page = page
        self.dados = dados
        self.tem_diferenca = tem_diferenca
        self.pagina_atual = 0
        self.total_paginas = max(1, -(-len(dados) // self.LINHAS_POR_PAGINA))

        cor_cab = ft.Colors.RED_300 if tem_diferenca else ft.Colors.GREEN_700

        self.tabela = ft.DataTable(
            border=ft.border.all(1, ft.Colors.GREY_300),
            border_radius=8,
            vertical_lines=ft.BorderSide(1, ft.Colors.GREY_200),
            horizontal_lines=ft.BorderSide(1, ft.Colors.GREY_200),
            heading_row_color=cor_cab,
            heading_row_height=48,
            columns=[
                ft.DataColumn(
                    ft.Text(
                        "NF",
                        color=ft.Colors.WHITE,
                        weight=ft.FontWeight.BOLD
                    )
                ),
                ft.DataColumn(
                    ft.Text(
                        "Crédito R$",
                        color=ft.Colors.WHITE,
                        weight=ft.FontWeight.BOLD
                    ),
                    numeric=True,
                ),
                ft.DataColumn(
                    ft.Text(
                        "Débito R$",
                        color=ft.Colors.WHITE,
                        weight=ft.FontWeight.BOLD
                    ),
                    numeric=True,
                ),
                ft.DataColumn(
                    ft.Text(
                        "Diferença R$",
                        color=ft.Colors.WHITE,
                        weight=ft.FontWeight.BOLD
                    ),
                    numeric=True,
                ),
            ],
            rows=[],
        )

        self.txt_pagina = ft.Text("", size=13, color=ft.Colors.BLACK)

        self.btn_anterior = ft.IconButton(
            icon=ft.Icons.CHEVRON_LEFT,
            on_click=self.pagina_anterior,
            disabled=True,
        )
        self.btn_proximo = ft.IconButton(
            icon=ft.Icons.CHEVRON_RIGHT,
            on_click=self.proxima_pagina,
            disabled=self.total_paginas <= 1,
        )

        self._renderizar_pagina()

    def _renderizar_pagina(self):
        inicio = self.pagina_atual * self.LINHAS_POR_PAGINA
        fim = inicio + self.LINHAS_POR_PAGINA
        fatia = self.dados[inicio:fim]

        linhas = []
        for item in fatia:
            dif = item['diferenca']
            cor_dif = ft.Colors.RED_600 if dif != 0 else ft.Colors.GREEN_600
            linhas.append(
                ft.DataRow(cells=[
                    ft.DataCell(
                        ft.Text(
                            item['nf'],
                            color=ft.Colors.GREY_900,
                            weight=ft.FontWeight.BOLD,
                        )
                    ),
                    ft.DataCell(
                        ft.Text(
                            formatar_valor(item['credito']),
                            color=ft.Colors.GREY_900,
                            weight=ft.FontWeight.BOLD,
                        )
                    ),
                    ft.DataCell(
                        ft.Text(
                            formatar_valor(item['debito']),
                            color=ft.Colors.GREY_900,
                            weight=ft.FontWeight.BOLD,
                        )
                    ),
                    ft.DataCell(
                        ft.Text(
                            formatar_valor(dif),
                            color=cor_dif,
                            weight=ft.FontWeight.BOLD,
                        )
                    ),
                ])
            )

        self.tabela.rows = linhas
        self.txt_pagina.value = f"Página {self.pagina_atual + 1} de {self.total_paginas}  ({len(self.dados)} registros)"
        self.btn_anterior.disabled = self.pagina_atual == 0
        self.btn_proximo.disabled = self.pagina_atual >= self.total_paginas - 1

    def proxima_pagina(self, e):
        if self.pagina_atual < self.total_paginas - 1:
            self.pagina_atual += 1
            self._renderizar_pagina()
            self.page.update()

    def pagina_anterior(self, e):
        if self.pagina_atual > 0:
            self.pagina_atual -= 1
            self._renderizar_pagina()
            self.page.update()

    def build(self) -> ft.Column:
        return ft.Column(
            controls=[
                ft.Row(
                    controls=[
                        self.btn_anterior,
                        self.txt_pagina,
                        self.btn_proximo,
                    ],
                    alignment=ft.MainAxisAlignment.CENTER,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
                ft.ListView(
                    controls=[self.tabela],
                    expand=True,
                    auto_scroll=False,
                ),
            ],
            expand=True,
            spacing=8,
        )


# ── View principal ───────────────────────────────────────────────────────────

class AdiantamentoView:

    def __init__(self, page: ft.Page):
        self.page = page
        self.arquivo_selecionado = None

        self.file_picker = ft.FilePicker(on_result=self.on_arquivo_selecionado)

        self.txt_arquivo = ft.Text(
            "Nenhum arquivo selecionado.",
            italic=True,
            color=ft.Colors.BLACK,
            size=13,
        )

        self.status = ft.Text("", size=13)

        self._com_diferenca = []
        self._consolidado = []

        self.btn_exportar = ft.ElevatedButton(
            text="Exportar Excel",
            icon=ft.Icons.TABLE_VIEW,
            bgcolor=ft.Colors.GREEN_700,
            color=ft.Colors.WHITE,
            disabled=True,
            on_click=self.exportar,
        )
        self.container_diferenca = ft.Container(expand=True)
        self.container_consolidado = ft.Container(expand=True)

        self.tabs = ft.Tabs(
            selected_index=0,
            animation_duration=200,
            expand=True,
            tabs=[
                ft.Tab(
                    text="Com Diferença",
                    icon=ft.Icons.WARNING_AMBER_ROUNDED,
                    content=ft.Container(
                        content=self.container_diferenca,
                        padding=16,
                        expand=True,
                    ),
                ),
                ft.Tab(
                    text="Consolidado",
                    icon=ft.Icons.CHECK_CIRCLE_OUTLINE,
                    content=ft.Container(
                        content=self.container_consolidado,
                        padding=16,
                        expand=True,
                    ),
                ),
            ],
        )

    def on_arquivo_selecionado(self, e: ft.FilePickerResultEvent):
        if not e.files:
            self.txt_arquivo.value = "Nenhum arquivo selecionado."
            self.arquivo_selecionado = None
            self.page.update()
            return

        self.arquivo_selecionado = e.files[0].path
        self.txt_arquivo.value = f"Arquivo: {e.files[0].name}"
        self.status.value = "⏳ Processando..."
        self.status.color = ft.Colors.BLUE_600
        self.page.update()

        threading.Thread(target=self._executar_processamento, daemon=True).start()

    def _executar_processamento(self):
        try:
            com_diferenca, consolidado = processar_pdf(self.arquivo_selecionado)
            self._com_diferenca = com_diferenca
            self._consolidado = consolidado

            self.container_diferenca.content = (
                TabelaPaginada(self.page, com_diferenca, tem_diferenca=True).build()
                if com_diferenca
                else ft.Text(
                    "Nenhuma nota com diferença encontrada.",
                    italic=True,
                    color=ft.Colors.BLACK,
                )
            )

            self.container_consolidado.content = (
                TabelaPaginada(self.page, consolidado, tem_diferenca=False).build()
                if consolidado
                else ft.Text(
                    "Nenhuma nota consolidada encontrada.",
                    italic=True,
                    color=ft.Colors.BLACK,
                )
            )

            self.status.value = (
                f"✅ Processado — {len(com_diferenca)} nota(s) com diferença, "
                f"{len(consolidado)} nota(s) consolidada(s)."
            )
            self.status.color = ft.Colors.GREEN_700
            self.btn_exportar.disabled = False

        except Exception as ex:
            self.status.value = f"❌ Erro ao processar: {ex}"
            self.status.color = ft.Colors.RED_400
            self.btn_exportar.disabled = True

        self.page.update()

    def abrir_seletor(self, e):
        self.file_picker.pick_files(
            dialog_title="Selecionar PDF de Adiantamentos",
            allowed_extensions=["pdf"],
            allow_multiple=False,
        )

    def exportar(self, e):
        try:
            destino = exportar_excel(
                self.arquivo_selecionado,
                self._com_diferenca,
                self._consolidado,
            )
            self.status.value = f"✅ Excel salvo em: {destino}"
            self.status.color = ft.Colors.GREEN_700
        except Exception as ex:
            self.status.value = f"❌ Erro ao exportar: {ex}"
            self.status.color = ft.Colors.RED_600
        self.page.update()

    def build(self):
        return ft.Column(
            controls=[
                ft.Container(
                    content=ft.Column(
                        controls=[
                            ft.Text(
                                "Importar PDF",
                                size=16,
                                weight=ft.FontWeight.BOLD,
                                color=ft.Colors.BLUE_700,
                            ),
                            ft.Row(
                                controls=[
                                    ft.ElevatedButton(
                                        text="Selecionar PDF",
                                        icon=ft.Icons.UPLOAD_FILE,
                                        bgcolor=ft.Colors.BLUE_600,
                                        on_click=self.abrir_seletor,
                                    ),
                                    self.btn_exportar,
                                    self.txt_arquivo,
                                ],
                                spacing=16,
                                vertical_alignment=ft.CrossAxisAlignment.CENTER,
                            ),
                            self.status,
                        ],
                        spacing=10,
                    ),
                    padding=20,
                    border=ft.border.all(1, ft.Colors.GREY_300),
                    border_radius=8,
                    bgcolor=ft.Colors.WHITE,
                ),
                ft.Container(
                    content=self.tabs,
                    expand=True,
                    border=ft.border.all(1, ft.Colors.GREY_300),
                    border_radius=8,
                    bgcolor=ft.Colors.WHITE,
                ),
            ],
            spacing=16,
            expand=True,
        )
