import os
import re
import struct
import threading
import flet as ft
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

try:
    import olefile
    import xlrd
    _XLS_SUPPORT = True
except ImportError:
    _XLS_SUPPORT = False

# ─── Regex ────────────────────────────────────────────────────────────────────
# 1. NF/NFe + múltiplas separadas por /  (ex: "NF 544/26")
_RE_NF_MULT = re.compile(r'NFe?\s+(\d{2,6}(?:/\d{2,6})+)', re.I)
# 2. NF/NFe + número único — captura APENAS o primeiro (ex: "NF 64982 01/03" → 64982)
_RE_NF_SINGLE = re.compile(r'NFe?\s+(\d{2,6})', re.I)
# 3. Fallback sem keyword — primeiro número standalone (ex: "VENDAS NESTA DATA 26 COOP")
_RE_NF_FALLBACK = re.compile(r'(?<!\d)(\d{2,6})(?!\d)')


# ─── Helpers ──────────────────────────────────────────────────────────────────

def formatar_valor(valor: float) -> str:
    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _to_float(v) -> float:
    try:
        return float(v)
    except Exception:
        return 0.0


def _localizar_header(df_raw: pd.DataFrame) -> int:
    """Retorna o índice da linha que contém o cabeçalho real (coluna 'Histórico')."""
    for i, row in df_raw.iterrows():
        if row.astype(str).str.contains('Hist', case=False, na=False).any():
            return i
    raise ValueError("Cabeçalho com coluna 'Histórico' não encontrado no arquivo.")


def _col(df: pd.DataFrame, *parciais) -> str:
    """Retorna o nome da primeira coluna que contém algum dos termos parciais."""
    for p in parciais:
        for c in df.columns:
            if p.lower() in str(c).lower():
                return c
    raise KeyError(f"Coluna não encontrada para: {parciais}")


# ─── Parse ────────────────────────────────────────────────────────────────────

def _abrir_xls(caminho: str) -> pd.ExcelFile:
    """
    Abre um .xls sem LibreOffice.

    Alguns sistemas ERP geram arquivos .xls com o registro BOUNDSHEET apontando
    para um offset incorreto dentro do stream BIFF8. Isso faz xlrd (e qualquer
    outra biblioteca) falhar ao tentar ler a worksheet.

    A correção é feita em memória:
        1. Extrai o stream 'Workbook' via olefile (ignora o container OLE truncado)
        2. Localiza todos os BOF de Worksheet no stream (record 0x0809 tipo 0x0010)
        3. Localiza todos os BOUNDSHEET (0x0085) e corrige os offsets
        4. Passa o stream corrigido para xlrd → pandas
    """
    ole = olefile.OleFileIO(caminho)
    raw = bytearray(ole.openstream('Workbook').read())

    # Mapeia posições reais dos BOF de Worksheet
    bof_positions = []
    pos = 0
    while pos < len(raw) - 4:
        rtype, rlen = struct.unpack_from('<HH', raw, pos)
        if rtype == 0x0809 and rlen >= 4:
            ver = struct.unpack_from('<H', raw, pos + 4)[0]
            kind = struct.unpack_from('<H', raw, pos + 6)[0] if rlen >= 6 else 0
            if kind == 0x0010:  # Worksheet
                bof_positions.append(pos)
        pos += 4 + rlen

    # Corrige os BOUNDSHEET na ordem em que aparecem
    bsheet_idx = 0
    pos = 0
    while pos < len(raw) - 4 and bsheet_idx < len(bof_positions):
        rtype, rlen = struct.unpack_from('<HH', raw, pos)
        if rtype == 0x0085 and rlen >= 4:
            struct.pack_into('<I', raw, pos + 4, bof_positions[bsheet_idx])
            bsheet_idx += 1
        pos += 4 + rlen

    wb_xlrd = xlrd.open_workbook(file_contents=bytes(raw))
    # Reconstrói como ExcelFile via BytesIO em formato xlsx temporário
    # (mais simples: converte xlrd → dicts → pandas direto)
    dfs = {}
    for sh_name in wb_xlrd.sheet_names():
        sh = wb_xlrd.sheet_by_name(sh_name)
        rows = [sh.row_values(r) for r in range(sh.nrows)]
        dfs[sh_name] = pd.DataFrame(rows)
    return _DictExcelFile(dfs)


class _DictExcelFile:
    """Wrapper mínimo que imita pd.ExcelFile para dicts de DataFrames."""
    def __init__(self, sheets: dict):
        self._sheets = sheets
        self.sheet_names = list(sheets.keys())

    def parse(self, sheet_name, header=None, **kwargs):
        df = self._sheets[sheet_name].copy()
        if header is not None:
            df.columns = df.iloc[header]
            df = df.iloc[header + 1:].reset_index(drop=True)
        return df


def _parsear_excel(caminho: str) -> dict:
    """
    Lê todas as abas do arquivo e extrai pares (NF → crédito/débito).

    Regras:
        • Crédito > 0 → acumula como crédito da NF
        • Débito  > 0 → acumula como débito da NF
        • NF extraída do Histórico: primeiro número de 4–6 dígitos encontrado na linha
    """
    # .xls legado: corrige BOUNDSHEET corrompido em memória e passa para xlrd/pandas
    ext = os.path.splitext(caminho)[1].lower()
    if ext == '.xls':
        if not _XLS_SUPPORT:
            raise ImportError("Instale 'olefile' e 'xlrd' para abrir arquivos .xls.")
        xl = _abrir_xls(caminho)
    else:
        xl = pd.ExcelFile(caminho)
    dados: dict = {}
    pendentes_mult: list = []

    for sheet in xl.sheet_names:
        # _DictExcelFile (xls) usa .parse(); pd.ExcelFile (xlsx) usa pd.read_excel
        if isinstance(xl, _DictExcelFile):
            df_raw = xl.parse(sheet, header=None).astype(str)
        else:
            df_raw = pd.read_excel(caminho, sheet_name=sheet, header=None, dtype=str)
        try:
            hrow = _localizar_header(df_raw)
        except ValueError:
            continue  # aba sem dados reconhecíveis, pula

        if isinstance(xl, _DictExcelFile):
            df = xl.parse(sheet, header=hrow)
        else:
            df = pd.read_excel(caminho, sheet_name=sheet, header=hrow)

        try:
            hist = _col(df, 'hist')
            deb = _col(df, 'ébit', 'ebit')
            cred = _col(df, 'réd',  'red')
        except KeyError:
            continue  # aba sem as colunas esperadas, pula

        for _, row in df.iterrows():
            h = str(row[hist]) if pd.notna(row.get(hist)) else ''
            if not h or h.lower() == 'nan':
                continue

            deb_ = _to_float(row.get(deb))
            cred_ = _to_float(row.get(cred))

            if deb_ == 0 and cred_ == 0:
                continue

            # 1. Múltiplas NFs separadas por / : "NF 544/26", "NF 30/565"
            mult = _RE_NF_MULT.search(h)
            if mult:
                nfs = re.findall(r'\d{2,6}', mult.group(0))
                pendentes_mult.append({'nfs': nfs, 'cred': cred_, 'deb': deb_})
                continue

            # 2. NF única após keyword NF/NFe (ignora parcelas "01/03" depois)
            m = _RE_NF_SINGLE.search(h)
            if not m:
                # 3. Fallback: históricos sem keyword (ex: "VENDAS NESTA DATA 26 COOP")
                m = _RE_NF_FALLBACK.search(h)
            if not m:
                continue
            nf = m.group(1)
            dados.setdefault(nf, {'credito': 0.0, 'debito': 0.0})
            if cred_ > 0: dados[nf]['credito'] += cred_
            if deb_ > 0: dados[nf]['debito'] += deb_

    # Segunda passagem: distribui créditos múltiplos proporcionalmente ao débito
    for p in pendentes_mult:
        nfs = p['nfs']
        cred_ = p['cred']
        deb_ = p['deb']

        if cred_ > 0:
            # Proporção: cada NF recebe crédito proporcional ao seu débito acumulado
            total_deb = sum(dados.get(nf, {}).get('debito', 0.0) for nf in nfs)
            for nf in nfs:
                dados.setdefault(nf, {'credito': 0.0, 'debito': 0.0})
                if total_deb > 0:
                    proporcao = dados[nf]['debito'] / total_deb
                    dados[nf]['credito'] += round(cred_ * proporcao, 2)
                else:
                    dados[nf]['credito'] += round(cred_ / len(nfs), 2)

        if deb_ > 0:
            total_cred = sum(dados.get(nf, {}).get('credito', 0.0) for nf in nfs)
            for nf in nfs:
                dados.setdefault(nf, {'credito': 0.0, 'debito': 0.0})
                if total_cred > 0:
                    proporcao = dados[nf]['credito'] / total_cred
                    dados[nf]['debito'] += round(deb_ * proporcao, 2)
                else:
                    dados[nf]['debito'] += round(deb_ / len(nfs), 2)

    return dados


def _calcular_resultado(dados: dict):
    com_diferenca, consolidado = [], []
    for nf in sorted(dados, key=lambda x: int(x) if x.isdigit() else 0):
        c = round(dados[nf]['credito'], 2)
        d = round(dados[nf]['debito'],  2)
        dif = round(c - d, 2)
        item = {'nf': nf, 'credito': c, 'debito': d, 'diferenca': dif}
        if dif != 0:
            com_diferenca.append(item)
        elif c > 0:
            consolidado.append(item)
    return com_diferenca, consolidado


def processar_arquivo(caminho: str):
    if not os.path.exists(caminho):
        return [], []
    dados = _parsear_excel(caminho)
    return _calcular_resultado(dados)


# ─── Exportação Excel ─────────────────────────────────────────────────────────

def exportar_excel(caminho_origem: str, com_diferenca: list, consolidado: list) -> str:
    wb = openpyxl.Workbook()
    cab_font = Font(bold=True, color="FFFFFF")
    centro = Alignment(horizontal="center")
    colunas = ["NF", "Crédito R$", "Débito R$", "Diferença R$"]

    for titulo, dados, cor in [
        ("Com Diferença", com_diferenca, "C62828"),
        ("Consolidado",   consolidado,   "2E7D32"),
    ]:
        ws = wb.create_sheet(titulo)
        fill = PatternFill("solid", fgColor=cor)
        for col, cab in enumerate(colunas, 1):
            cell = ws.cell(1, col, cab)
            cell.font, cell.fill, cell.alignment = cab_font, fill, centro
        for linha, item in enumerate(dados, 2):
            ws.cell(linha, 1, item['nf'])
            ws.cell(linha, 2, item['credito'])
            ws.cell(linha, 3, item['debito'])
            ws.cell(linha, 4, item['diferenca'])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    del wb["Sheet"]
    destino = f"{os.path.splitext(caminho_origem)[0]}_consolidado.xlsx"
    wb.save(destino)
    return destino


# ─── Tabela paginada ──────────────────────────────────────────────────────────

class TabelaPaginada:
    LPP = 50

    def __init__(self, page: ft.Page, dados: list, tem_diferenca: bool):
        self.page = page
        self.dados = dados
        self.pagina_atual = 0
        self.total_pags = max(1, -(-len(dados) // self.LPP))
        cor = ft.Colors.RED_300 if tem_diferenca else ft.Colors.GREEN_700

        self.tabela = ft.DataTable(
            border=ft.border.all(1, ft.Colors.GREY_300),
            border_radius=8,
            vertical_lines=ft.BorderSide(1, ft.Colors.GREY_200),
            horizontal_lines=ft.BorderSide(1, ft.Colors.GREY_200),
            heading_row_color=cor,
            heading_row_height=48,
            columns=[
                ft.DataColumn(ft.Text("NF", color=ft.Colors.WHITE, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Crédito R$", color=ft.Colors.WHITE, weight=ft.FontWeight.BOLD), numeric=True),
                ft.DataColumn(ft.Text("Débito R$", color=ft.Colors.WHITE, weight=ft.FontWeight.BOLD), numeric=True),
                ft.DataColumn(ft.Text("Diferença R$", color=ft.Colors.WHITE, weight=ft.FontWeight.BOLD), numeric=True),
            ],
            rows=[],
        )
        self.txt_pag = ft.Text("", size=13, color=ft.Colors.BLACK)
        self.btn_prev = ft.IconButton(ft.Icons.CHEVRON_LEFT,  on_click=self._anterior, disabled=True)
        self.btn_next = ft.IconButton(ft.Icons.CHEVRON_RIGHT, on_click=self._proximo,  disabled=self.total_pags <= 1)
        self._render()

    def _render(self):
        inicio = self.pagina_atual * self.LPP
        linhas = []
        for item in self.dados[inicio: inicio + self.LPP]:
            dif = item['diferenca']
            cor = ft.Colors.RED_600 if dif != 0 else ft.Colors.GREEN_600
            linhas.append(ft.DataRow(cells=[
                ft.DataCell(ft.Text(item['nf'], color=ft.Colors.GREY_900, weight=ft.FontWeight.BOLD)),
                ft.DataCell(ft.Text(formatar_valor(item['credito']), color=ft.Colors.GREY_900, weight=ft.FontWeight.BOLD)),
                ft.DataCell(ft.Text(formatar_valor(item['debito']),  color=ft.Colors.GREY_900, weight=ft.FontWeight.BOLD)),
                ft.DataCell(ft.Text(formatar_valor(dif), color=cor, weight=ft.FontWeight.BOLD)),
            ]))
        self.tabela.rows = linhas
        self.txt_pag.value = (
            f"Página {self.pagina_atual + 1} de {self.total_pags}  ({len(self.dados)} registros)"
        )
        self.btn_prev.disabled = self.pagina_atual == 0
        self.btn_next.disabled = self.pagina_atual >= self.total_pags - 1

    def _proximo(self, e):
        if self.pagina_atual < self.total_pags - 1:
            self.pagina_atual += 1
            self._render()
            self.page.update()

    def _anterior(self, e):
        if self.pagina_atual > 0:
            self.pagina_atual -= 1
            self._render()
            self.page.update()

    def build(self) -> ft.Column:
        return ft.Column(
            controls=[
                ft.Row(
                    controls=[self.btn_prev, self.txt_pag, self.btn_next],
                    alignment=ft.MainAxisAlignment.CENTER,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
                ft.ListView(controls=[self.tabela], expand=True, auto_scroll=False),
            ],
            expand=True, spacing=8,
        )


# ─── View principal ───────────────────────────────────────────────────────────

class AdiantamentoView:

    def __init__(self, page: ft.Page):
        self.page = page
        self.arquivo_selecionado = None
        self._com_diferenca = []
        self._consolidado = []

        self.file_picker = ft.FilePicker(on_result=self.on_arquivo)
        self.txt_arquivo = ft.Text(
            "Nenhum arquivo selecionado.", italic=True,
            color=ft.Colors.BLACK, size=13,
        )
        self.status = ft.Text("", size=13)
        self.btn_exportar = ft.ElevatedButton(
            "Exportar Excel", icon=ft.Icons.TABLE_VIEW,
            bgcolor=ft.Colors.GREEN_700, color=ft.Colors.WHITE,
            disabled=True, on_click=self.exportar,
        )

        self.container_dif = ft.Container(expand=True)
        self.container_cons = ft.Container(expand=True)

        self.tabs = ft.Tabs(
            selected_index=0, animation_duration=200, expand=True,
            tabs=[
                ft.Tab(
                    text="Com Diferença",
                    icon=ft.Icons.WARNING_AMBER_ROUNDED,
                    content=ft.Container(
                        content=self.container_dif, padding=16, expand=True
                    ),
                ),
                ft.Tab(
                    text="Consolidado",
                    icon=ft.Icons.CHECK_CIRCLE_OUTLINE,
                    content=ft.Container(
                        content=self.container_cons, padding=16, expand=True
                    ),
                ),
            ],
        )

    def on_arquivo(self, e: ft.FilePickerResultEvent):
        if not e.files:
            self.txt_arquivo.value = "Nenhum arquivo selecionado."
            self.arquivo_selecionado = None
            self.page.update()
            return
        self.arquivo_selecionado = e.files[0].path
        self.txt_arquivo.value = f"Arquivo: {e.files[0].name}"
        self.status.value = "⏳ Processando..."
        self.status.color = ft.Colors.BLUE_600
        self.page.update()
        threading.Thread(target=self._processar, daemon=True).start()

    def _processar(self):
        try:
            com_dif, cons = processar_arquivo(self.arquivo_selecionado)
            self._com_diferenca = com_dif
            self._consolidado = cons

            self.container_dif.content = (
                TabelaPaginada(self.page, com_dif, True).build()
                if com_dif
                else ft.Text(
                    "Nenhuma nota com diferença encontrada.",
                    italic=True, color=ft.Colors.BLACK,
                )
            )
            self.container_cons.content = (
                TabelaPaginada(self.page, cons, False).build()
                if cons
                else ft.Text(
                    "Nenhuma nota consolidada encontrada.",
                    italic=True, color=ft.Colors.BLACK,
                )
            )
            self.status.value = (
                f"✅ {len(com_dif)} nota(s) com diferença, "
                f"{len(cons)} nota(s) consolidada(s)."
            )
            self.status.color = ft.Colors.GREEN_700
            self.btn_exportar.disabled = False

        except Exception as ex:
            self.status.value = f"❌ Erro ao processar: {ex}"
            self.status.color = ft.Colors.RED_400
            self.btn_exportar.disabled = True

        self.page.update()

    def abrir_seletor(self, e):
        self.file_picker.pick_files(
            dialog_title="Selecionar planilha",
            allowed_extensions=["xls", "xlsx", "xlsm"],
            allow_multiple=False,
        )

    def exportar(self, e):
        try:
            destino = exportar_excel(
                self.arquivo_selecionado,
                self._com_diferenca,
                self._consolidado,
            )
            self.status.value = f"✅ Excel salvo em: {destino}"
            self.status.color = ft.Colors.GREEN_700
        except Exception as ex:
            self.status.value = f"❌ Erro ao exportar: {ex}"
            self.status.color = ft.Colors.RED_600
        self.page.update()

    def build(self):
        return ft.Column(
            controls=[
                ft.Container(
                    content=ft.Column(
                        controls=[
                            ft.Text(
                                "Importar Planilha",
                                size=16, weight=ft.FontWeight.BOLD,
                                color=ft.Colors.BLUE_700,
                            ),
                            ft.Row(
                                controls=[
                                    ft.ElevatedButton(
                                        "Selecionar Arquivo",
                                        icon=ft.Icons.UPLOAD_FILE,
                                        bgcolor=ft.Colors.BLUE_600,
                                        on_click=self.abrir_seletor,
                                    ),
                                    self.btn_exportar,
                                    self.txt_arquivo,
                                ],
                                spacing=12,
                                vertical_alignment=ft.CrossAxisAlignment.CENTER,
                            ),
                            self.status,
                        ],
                        spacing=10,
                    ),
                    padding=20,
                    border=ft.border.all(1, ft.Colors.GREY_300),
                    border_radius=8,
                    bgcolor=ft.Colors.WHITE,
                ),
                ft.Container(
                    content=self.tabs,
                    expand=True,
                    border=ft.border.all(1, ft.Colors.GREY_300),
                    border_radius=8,
                    bgcolor=ft.Colors.WHITE,
                ),
            ],
            spacing=16,
            expand=True,
        )
